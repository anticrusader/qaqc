#!/usr/bin/env python3
"""
Check Excel file structure to understand column layout
"""

import openpyxl

def check_excel_structure(excel_path):
    """Check the structure of Excel file at row 12"""
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        print(f"Checking Excel file: {excel_path}")
        print("Row 12 data:")
        
        for col in range(1, 15):  # Check first 14 columns
            cell_value = sheet.cell(row=12, column=col).value
            print(f"  Column {col}: '{cell_value}'")
        
        print("\nRow 11 data (headers?):")
        for col in range(1, 15):
            cell_value = sheet.cell(row=11, column=col).value
            print(f"  Column {col}: '{cell_value}'")
        
        workbook.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    # Check the first Excel file
    check_excel_structure("L01-H01D01-FOS-00-XX-MUP-AR-80050[T0].xlsx")