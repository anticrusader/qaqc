#!/usr/bin/env python3
"""
Excel-Based PDF Extractor

New Logic:
1. Read Document Number, Revision, and Title from Excel file (line 12)
2. Validate these values exist in PDF content
3. Check revision history table for latest_revision validation
4. Extract table_title from PDF content
5. No latest_reason extraction needed

Author: AI Assistant
Version: 4.0 Excel-Based
"""

import fitz
import pandas as pd
import re
from pathlib import Path
import logging
import openpyxl

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pdf_extraction_excel_based.log'),
        logging.StreamHandler()
    ]
)

def read_excel_data(excel_path, row_number=12):
    """
    Read Document Number, Revision, and Title from Excel file at specified row.
    
    Args:
        excel_path (str): Path to Excel file
        row_number (int): Row number to read (default: 12)
        
    Returns:
        dict: Contains document_number, revision, title or None if error
    """
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active  # Use active sheet
        
        # Read row 12 - Column 1: Document No, Column 2: Revision, Column 3: Title
        document_number = sheet.cell(row=row_number, column=1).value  # Document No
        revision = sheet.cell(row=row_number, column=2).value         # Revision  
        title = sheet.cell(row=row_number, column=3).value           # Title
        
        # Clean up values (remove None, strip whitespace, handle line breaks)
        document_number = str(document_number).strip() if document_number else ""
        
        # Handle revision formatting - add leading zero for single digits
        if revision is not None:
            revision_str = str(revision).strip()
            
            # Check if it's a single digit number (like 7) that needs to be formatted as 07
            if revision_str.isdigit() and len(revision_str) == 1:
                revision = f"0{revision_str}"  # Format as 07, 08, etc.
                logging.debug(f"Formatted single digit revision '{revision_str}' as '{revision}'")
            else:
                revision = revision_str
        else:
            revision = ""
        
        # Handle title with potential line breaks - replace line breaks with spaces
        if title:
            title = str(title).replace('\n', ' ').replace('\r', ' ').strip()
            # Clean up multiple spaces
            title = ' '.join(title.split())
        else:
            title = ""
        
        workbook.close()
        
        logging.info(f"Excel data read - Document: '{document_number}', Revision: '{revision}', Title: '{title}'")
        
        return {
            'document_number': document_number,
            'revision': revision,
            'title': title
        }
        
    except Exception as e:
        logging.error(f"Error reading Excel file {excel_path}: {e}")
        return None

def validate_text_in_pdf(page, text_to_find):
    """
    Check if specific text exists in PDF content with improved matching.
    
    Args:
        page: PyMuPDF page object
        text_to_find (str): Text to search for
        
    Returns:
        bool: True if text found, False otherwise
    """
    
    if not text_to_find:
        return False
    
    # Get all text from PDF
    text_dict = page.get_text("dict")
    
    # Collect all text content for comprehensive search
    all_pdf_text = ""
    
    # Search through all text blocks
    for block in text_dict["blocks"]:
        if "lines" not in block:
            continue
            
        for line in block["lines"]:
            for span in line["spans"]:
                span_text = span["text"].strip()
                all_pdf_text += " " + span_text
                
                # Check for exact match (case-insensitive)
                if text_to_find.lower() in span_text.lower():
                    logging.debug(f"Found exact text '{text_to_find}' in PDF span: '{span_text}'")
                    return True
    
    # Clean up all PDF text for comprehensive matching
    all_pdf_text = ' '.join(all_pdf_text.split()).lower()
    text_to_find_clean = ' '.join(text_to_find.split()).lower()
    
    # Check if the entire text exists in PDF (case-insensitive)
    if text_to_find_clean in all_pdf_text:
        logging.debug(f"Found complete text '{text_to_find}' in combined PDF content")
        return True
    
    # For long titles, check word-by-word matching
    if len(text_to_find) > 15:
        words_to_find = text_to_find_clean.split()
        words_found = 0
        
        for word in words_to_find:
            if len(word) > 2 and word in all_pdf_text:  # Skip very short words
                words_found += 1
        
        match_ratio = words_found / len(words_to_find) if words_to_find else 0
        
        if match_ratio >= 0.6:  # 60% of significant words match
            logging.debug(f"Found word-based match for '{text_to_find}' (ratio: {match_ratio:.2f})")
            return True
    
    # Try splitting the title and looking for key phrases
    # Common title patterns in architectural drawings
    key_phrases = []
    
    # Extract key architectural terms
    architectural_terms = [
        'mockup', 'mock-up', 'external', 'wall', 'systems', 'typical', 'facade', 'section', 'details',
        'mep', 'door', 'room', 'grms', 'layout', 'technical', 'project', 'information', 'cover', 'sheet',
        'pool', 'enlargement', 'plan', 'grading', 'drainage', 'piping', 'conduit', 'overall'
    ]
    
    # Check if key architectural terms from title exist in PDF
    title_words = text_to_find_clean.split()
    key_terms_found = 0
    key_terms_total = 0
    
    for word in title_words:
        if word in architectural_terms:
            key_terms_total += 1
            if word in all_pdf_text:
                key_terms_found += 1
    
    if key_terms_total > 0:
        key_term_ratio = key_terms_found / key_terms_total
        if key_term_ratio >= 0.5:  # 50% of key architectural terms found
            logging.debug(f"Found key terms match for '{text_to_find}' (key term ratio: {key_term_ratio:.2f})")
            return True
    
    logging.warning(f"Text '{text_to_find}' not found in PDF content")
    return False

def find_revision_in_history_table(page, revision_to_find):
    """
    Check if revision exists in the revision history table.
    
    Args:
        page: PyMuPDF page object
        revision_to_find (str): Revision code to search for
        
    Returns:
        bool: True if revision found in history table, False otherwise
    """
    
    if not revision_to_find:
        return False
    
    text_dict = page.get_text("dict")
    page_width = page.rect.width
    page_height = page.rect.height
    
    # Look in title block area where revision tables are typically located
    title_block_x_start = page_width * 0.6   # Right 40% of page
    title_block_y_start = page_height * 0.3   # Bottom 70% of page (expanded search)
    
    revision_found_in_table = False
    
    # Search in title block area first
    for block in text_dict["blocks"]:
        if "lines" not in block:
            continue
            
        for line in block["lines"]:
            for span in line["spans"]:
                text = span["text"].strip()
                bbox = span["bbox"]
                
                # Check if in title block area
                if bbox[0] >= title_block_x_start and bbox[1] >= title_block_y_start:
                    
                    # Look for exact revision match
                    if revision_to_find.upper() == text.upper():
                        logging.debug(f"Found revision '{revision_to_find}' in history table at position {bbox}")
                        revision_found_in_table = True
                        break
                    
                    # Also check if revision is part of a larger text block
                    if revision_to_find.upper() in text.upper():
                        # Additional validation - make sure it's not part of another word
                        pattern = r'\b' + re.escape(revision_to_find.upper()) + r'\b'
                        if re.search(pattern, text.upper()):
                            logging.debug(f"Found revision '{revision_to_find}' in history table text: '{text}'")
                            revision_found_in_table = True
                            break
    
    # If not found in title block, search entire page
    if not revision_found_in_table:
        logging.debug(f"Revision '{revision_to_find}' not found in title block, searching entire page")
        
        for block in text_dict["blocks"]:
            if "lines" not in block:
                continue
                
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    
                    # Look for revision in any text
                    if revision_to_find.upper() == text.upper():
                        logging.debug(f"Found revision '{revision_to_find}' in page content: '{text}'")
                        revision_found_in_table = True
                        break
                    
                    # Pattern matching for revisions in context
                    pattern = r'\b' + re.escape(revision_to_find.upper()) + r'\b'
                    if re.search(pattern, text.upper()):
                        logging.debug(f"Found revision '{revision_to_find}' in page text: '{text}'")
                        revision_found_in_table = True
                        break
    
    if revision_found_in_table:
        logging.info(f"Revision '{revision_to_find}' validated in PDF history table")
    else:
        logging.warning(f"Revision '{revision_to_find}' not found in PDF history table")
    
    return revision_found_in_table

def extract_table_title_from_pdf(page):
    """
    Extract table title - only the 5 standard phases.
    
    Args:
        page: PyMuPDF page object
        
    Returns:
        str: Table title from the 5 standard phases
    """
    
    text_dict = page.get_text("dict")
    
    # The 5 standard table titles in order of preference
    standard_table_titles = [
        'Concept Design',
        'Schematic Design',
        'Design Development', 
        'Construction Documents',
        'Construction Procurement'
    ]
    
    # Search for table titles in PDF content
    for block in text_dict["blocks"]:
        if "lines" not in block:
            continue
            
        for line in block["lines"]:
            for span in line["spans"]:
                text = span["text"].strip()
                
                # Check if text matches any of the 5 standard table titles
                for title in standard_table_titles:
                    if title.lower() in text.lower():
                        logging.debug(f"Table title found: {title}")
                        return title
    
    # Default to most common table title
    default_title = "Construction Procurement"
    logging.debug(f"No table title found, using default: {default_title}")
    return default_title

def process_pdf_with_excel(pdf_path):
    """
    Process PDF using Excel file as source of truth.
    
    Args:
        pdf_path (str): Path to PDF file
        
    Returns:
        dict: Extraction results
    """
    
    pdf_file = Path(pdf_path)
    filename = pdf_file.name
    
    # Construct Excel file path (same name as PDF but .xlsx extension)
    excel_path = pdf_file.with_suffix('.xlsx')
    
    logging.info(f"Processing PDF: {filename}")
    logging.info(f"Looking for Excel file: {excel_path}")
    
    try:
        # Check if Excel file exists
        if not excel_path.exists():
            return {
                'file_name': filename,
                'drawing_title': "",
                'drawing_number': "",
                'revision': "",
                'latest_revision': "",
                'latest_date': "",
                'latest_reason': "",  # Not needed but keeping for compatibility
                'table_title': "",
                'status': f"ERROR: Excel file not found: {excel_path.name}"
            }
        
        # Read data from Excel file
        excel_data = read_excel_data(str(excel_path))
        if not excel_data:
            return {
                'file_name': filename,
                'drawing_title': "",
                'drawing_number': "",
                'revision': "",
                'latest_revision': "",
                'latest_date': "",
                'latest_reason': "",
                'table_title': "",
                'status': "ERROR: Failed to read Excel data"
            }
        
        # Open PDF for validation
        doc = fitz.open(pdf_path)
        page = doc[0]  # Process first page
        
        # Validate Excel data against PDF content
        document_number_valid = validate_text_in_pdf(page, excel_data['document_number'])
        revision_valid = validate_text_in_pdf(page, excel_data['revision'])
        title_valid = validate_text_in_pdf(page, excel_data['title'])
        
        # Check if revision exists in history table
        revision_in_history = find_revision_in_history_table(page, excel_data['revision'])
        
        # Extract table title from PDF
        table_title = extract_table_title_from_pdf(page)
        
        doc.close()
        
        # Determine latest_revision based on history table validation
        latest_revision = excel_data['revision'] if revision_in_history else "NULL"
        
        # Create result
        result = {
            'file_name': filename,
            'drawing_title': excel_data['title'] if title_valid else "",
            'drawing_number': excel_data['document_number'] if document_number_valid else "",
            'revision': excel_data['revision'] if revision_valid else "",
            'latest_revision': latest_revision,
            'latest_date': "",  # Not extracted in this version
            'latest_reason': "",  # Not needed
            'table_title': table_title,
            'status': ""
        }
        
        # Determine status based on validation results
        validation_issues = []
        
        if not document_number_valid:
            validation_issues.append("Document number not found in PDF")
        if not revision_valid:
            validation_issues.append("Revision not found in PDF")
        if not title_valid:
            validation_issues.append("Title not found in PDF")
        if not revision_in_history:
            validation_issues.append("Revision not found in history table")
        
        if validation_issues:
            result['status'] = "FAILED - " + "; ".join(validation_issues)
        else:
            result['status'] = "SUCCESS"
        
        logging.info(f"Processing complete for {filename}: {result['status']}")
        return result
        
    except Exception as e:
        logging.error(f"Error processing {filename}: {e}")
        return {
            'file_name': filename,
            'drawing_title': "",
            'drawing_number': "",
            'revision': "",
            'latest_revision': "",
            'latest_date': "",
            'latest_reason': "",
            'table_title': "",
            'status': f"ERROR: {str(e)}"
        }

def process_all_pdfs_with_excel():
    """
    Process all PDF files using their corresponding Excel files.
    
    Returns:
        list: List of extraction results
    """
    
    # Find all PDF files in current directory
    pdf_files = list(Path('.').glob('*.pdf'))
    
    if not pdf_files:
        logging.warning("No PDF files found in current directory")
        return []
    
    logging.info(f"Found {len(pdf_files)} PDF files to process")
    
    results = []
    
    for pdf_file in pdf_files:
        result = process_pdf_with_excel(str(pdf_file))
        results.append(result)
    
    return results

def save_results_to_csv(results, output_filename="pdf_extraction_results_excel_based.csv"):
    """Save extraction results to CSV file with proper UTF-8 encoding and string preservation."""
    
    if not results:
        logging.warning("No results to save")
        return
    
    df = pd.DataFrame(results)
    
    # Ensure revision columns are treated as strings and add apostrophe prefix for Excel
    if 'revision' in df.columns:
        df['revision'] = df['revision'].apply(lambda x: f"'{x}" if pd.notna(x) and str(x) else x)
    if 'latest_revision' in df.columns:
        df['latest_revision'] = df['latest_revision'].apply(lambda x: f"'{x}" if pd.notna(x) and str(x) else x)
    
    # Save with UTF-8 BOM encoding for better compatibility with Excel and other programs
    df.to_csv(output_filename, index=False, encoding='utf-8-sig')
    logging.info(f"Results saved to: {output_filename} with UTF-8-sig encoding")

def print_detailed_results(results):
    """Print detailed results with validation status."""
    
    print("\n📊 EXCEL-BASED EXTRACTION RESULTS:")
    print("=" * 80)
    
    for result in results:
        print(f"\n📋 {result['file_name']}")
        print(f"  📝 Title: '{result['drawing_title']}'")
        print(f"  🔢 Number: '{result['drawing_number']}'")
        print(f"  📊 Revision: '{result['revision']}'")
        print(f"  📈 Latest Revision: '{result['latest_revision']}'")
        print(f"  📋 Table Title: '{result['table_title']}'")
        print(f"  ✅ Status: {result['status']}")
    
    # Summary statistics
    total_files = len(results)
    success_count = len([r for r in results if r['status'] == 'SUCCESS'])
    failed_count = len([r for r in results if 'FAILED' in r['status']])
    error_count = len([r for r in results if 'ERROR' in r['status']])
    
    print(f"\n📊 SUMMARY:")
    print(f"  📁 Total Files: {total_files}")
    print(f"  ✅ Successful: {success_count} ({success_count/total_files*100:.1f}%)")
    print(f"  ❌ Failed Validation: {failed_count} ({failed_count/total_files*100:.1f}%)")
    print(f"  🚫 Errors: {error_count} ({error_count/total_files*100:.1f}%)")

def main():
    """
    Main function - Excel-based PDF extraction and validation.
    """
    
    print("📊 STARTING EXCEL-BASED PDF EXTRACTION...")
    print("=" * 80)
    print("Logic:")
    print("1. Read Document Number, Revision, Title from Excel file (row 12)")
    print("2. Validate these values exist in PDF content")
    print("3. Check revision in history table for latest_revision")
    print("4. Extract table_title from PDF content")
    print("5. No latest_reason extraction")
    
    # Process all PDFs with their Excel files
    results = process_all_pdfs_with_excel()
    
    if not results:
        print("No PDF files found to process")
        return
    
    # Save results to CSV
    output_file = 'pdf_extraction_results_excel_based_fixed.csv'
    save_results_to_csv(results, output_file)
    
    # Print detailed results
    print_detailed_results(results)
    
    print(f"\n✅ Excel-based extraction complete! Results saved to: {output_file}")

if __name__ == "__main__":
    main()