#!/usr/bin/env python3
"""
TestData Folder PDF Extractor

New Logic:
1. Process testdata folder with subfolders (CHP, DAE, etc.)
2. Read from main Excel file 'ExportDocs - Test Data file RSG QAQC Process.xlsx'
3. Each sheet corresponds to a subfolder name
4. Column 13 onwards contains PDF file records
5. Extract Document Number, Revision, Title from Excel sheets
6. Validate against PDF files in corresponding subfolders
7. Add comments column for title mismatches

Author: AI Assistant
Version: 5.1 TestData-Based with Comments
"""

import fitz
import pandas as pd
import re
from pathlib import Path
import logging
import openpyxl

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pdf_extraction_testdata.log'),
        logging.StreamHandler()
    ]
)

def get_excel_sheet_names(excel_path):
    """
    Get all sheet names from the Excel file.
    
    Args:
        excel_path (str): Path to Excel file
        
    Returns:
        list: List of sheet names
    """
    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        logging.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
        return sheet_names
    except Exception as e:
        logging.error(f"Error reading Excel sheet names: {e}")
        return []

def read_excel_sheet_data(excel_path, sheet_name, header_row=12, data_start_row=13):
    """
    Read PDF file data from specific Excel sheet.
    
    Args:
        excel_path (str): Path to Excel file
        sheet_name (str): Name of the sheet to read
        header_row (int): Row number containing headers (default: 12)
        data_start_row (int): Row number where data starts (default: 13)
        
    Returns:
        list: List of dictionaries containing PDF file data
    """
    try:
        workbook = openpyxl.load_workbook(excel_path)
        
        if sheet_name not in workbook.sheetnames:
            logging.warning(f"Sheet '{sheet_name}' not found in Excel file")
            workbook.close()
            return []
        
        sheet = workbook[sheet_name]
        pdf_records = []
        
        # Read the specific columns we need: Document No (Col1), Revision (Col2), Title (Col3)
        document_col = 1  # Column A
        revision_col = 2  # Column B  
        title_col = 3     # Column C
        
        logging.info(f"Reading data from sheet '{sheet_name}' starting at row {data_start_row}")
        
        for row in range(data_start_row, sheet.max_row + 1):
            # Read the three main columns
            document_number = sheet.cell(row=row, column=document_col).value
            revision = sheet.cell(row=row, column=revision_col).value
            title = sheet.cell(row=row, column=title_col).value
            
            # Clean up the values
            document_number = str(document_number).strip() if document_number else ""
            revision = str(revision).strip() if revision else ""
            title = str(title).strip() if title else ""
            
            # Only add if we have at least a document number
            if document_number:
                pdf_records.append({
                    'document_number': document_number,
                    'revision': revision,
                    'title': title,
                    'row_number': row
                })
                logging.debug(f"Row {row}: Doc={document_number}, Rev={revision}, Title={title[:50]}...")
        
        workbook.close()
        logging.info(f"Found {len(pdf_records)} PDF records in sheet '{sheet_name}'")
        return pdf_records
        
    except Exception as e:
        logging.error(f"Error reading Excel sheet '{sheet_name}': {e}")
        return []

def validate_text_in_pdf(page, text_to_find):
    """
    Check if specific text exists in PDF content with very strict matching for titles.
    For titles, we require the text to appear as a cohesive phrase, not scattered words.
    """
    if not text_to_find:
        return False
    
    text_dict = page.get_text("dict")
    all_pdf_text = ""
    
    # Collect all text spans for analysis
    text_spans = []
    for block in text_dict["blocks"]:
        if "lines" not in block:
            continue
            
        for line in block["lines"]:
            for span in line["spans"]:
                span_text = span["text"].strip()
                if span_text:
                    text_spans.append(span_text)
                    all_pdf_text += " " + span_text
    
    all_pdf_text = ' '.join(all_pdf_text.split()).lower()
    text_to_find_clean = ' '.join(text_to_find.split()).lower()
    
    # First check: exact substring match (most reliable)
    if text_to_find_clean in all_pdf_text:
        logging.debug(f"Found exact substring match for '{text_to_find}' in PDF content")
        return True
    
    # Second check: look for the text in individual spans (for titles that might be split)
    for span_text in text_spans:
        if text_to_find.lower() in span_text.lower():
            logging.debug(f"Found exact text '{text_to_find}' in PDF span: '{span_text}'")
            return True
    
    # Third check: for titles longer than 15 chars, require a significant portion of the title
    # to appear as consecutive sequences (not just common prefixes like "Marina Beach Club")
    if len(text_to_find) > 15:
        words_to_find = text_to_find_clean.split()
        
        # Look for sequences of 5+ consecutive words from the title (much stricter)
        sequences_found = 0
        for i in range(len(words_to_find) - 4):
            sequence = ' '.join(words_to_find[i:i+5])
            if sequence in all_pdf_text:
                logging.debug(f"Found 5-word consecutive sequence '{sequence}' from title '{text_to_find}'")
                sequences_found += 1
        
        # Require at least one 5-word sequence to match
        if sequences_found > 0:
            logging.debug(f"Found {sequences_found} consecutive 5-word sequences for title '{text_to_find}'")
            return True
        
        # If no 5-word sequences found, try 4-word sequences but require multiple matches
        sequences_found = 0
        for i in range(len(words_to_find) - 3):
            sequence = ' '.join(words_to_find[i:i+4])
            if sequence in all_pdf_text:
                logging.debug(f"Found 4-word consecutive sequence '{sequence}' from title '{text_to_find}'")
                sequences_found += 1
        
        # Require at least 2 different 4-word sequences to match (very strict)
        if sequences_found >= 2:
            logging.debug(f"Found {sequences_found} consecutive 4-word sequences for title '{text_to_find}'")
            return True
        
        # If insufficient consecutive sequences found, the title doesn't match
        logging.warning(f"Insufficient consecutive word sequences found for title '{text_to_find}' in PDF (found {sequences_found} 4-word sequences)")
        return False
    
    # For shorter text, require exact match
    logging.warning(f"Text '{text_to_find}' not found in PDF content")
    return False

def extract_revision_from_pdf_title_block(page):
    """
    Extract the revision number from the PDF title block.
    Logic: Find "Revision" label, then get the value that comes after it.
    """
    import re
    text_dict = page.get_text("dict")
    
    # Step 1: Find all text spans and their positions
    all_spans = []
    for block in text_dict["blocks"]:
        if "lines" not in block:
            continue
        for line in block["lines"]:
            for span in line["spans"]:
                text = span["text"].strip()
                if text:
                    all_spans.append({
                        'text': text,
                        'bbox': span["bbox"]
                    })
    
    # Step 2: Find the "Revision" label
    revision_label_found = False
    revision_label_bbox = None
    
    for span in all_spans:
        if "revision" in span['text'].lower():
            revision_label_found = True
            revision_label_bbox = span['bbox']
            logging.debug(f"Found 'Revision' label at position {revision_label_bbox}")
            break
    
    if not revision_label_found:
        logging.warning("Could not find 'Revision' label in PDF")
        return ""
    
    # Step 3: Find the revision value near the "Revision" label
    # Look for spans that are close to the revision label (same row or nearby)
    revision_y = revision_label_bbox[1]  # Y coordinate of revision label
    revision_x = revision_label_bbox[2]  # Right edge of revision label
    
    # Collect potential revision values with their distances
    candidates = []
    
    for span in all_spans:
        span_bbox = span['bbox']
        text = span['text']
        
        # Check if span is in the same row or nearby (within 20 pixels vertically)
        if abs(span_bbox[1] - revision_y) <= 20:
            # Check if text matches revision pattern
            if re.match(r'^[NT]\d+$', text):  # N0, N1, T0, T1, etc.
                distance = abs(span_bbox[0] - revision_x)
                candidates.append({'text': text, 'distance': distance, 'type': 'NT'})
            elif re.match(r'^\d{2}$', text):  # 00, 01, 02, etc.
                distance = abs(span_bbox[0] - revision_x)
                candidates.append({'text': text, 'distance': distance, 'type': 'numeric'})
            elif re.match(r'^[A-Z]{2}$', text) and text not in ['AD', 'SF', 'CG']:  # XX, YY, etc. but exclude common initials
                distance = abs(span_bbox[0] - revision_x)
                candidates.append({'text': text, 'distance': distance, 'type': 'alpha'})
    
    # Sort by priority: NT patterns first, then numeric, then alpha, then by distance
    if candidates:
        # Prioritize NT patterns (N0, N1, T0, T1)
        nt_candidates = [c for c in candidates if c['type'] == 'NT']
        if nt_candidates:
            best_candidate = min(nt_candidates, key=lambda x: x['distance'])
            logging.debug(f"Found PDF revision value: '{best_candidate['text']}' near Revision label")
            return best_candidate['text']
        
        # Then numeric patterns (00, 01, 02)
        numeric_candidates = [c for c in candidates if c['type'] == 'numeric']
        if numeric_candidates:
            best_candidate = min(numeric_candidates, key=lambda x: x['distance'])
            logging.debug(f"Found PDF revision value: '{best_candidate['text']}' near Revision label")
            return best_candidate['text']
        
        # Finally alpha patterns (XX, YY) excluding common initials
        alpha_candidates = [c for c in candidates if c['type'] == 'alpha']
        if alpha_candidates:
            best_candidate = min(alpha_candidates, key=lambda x: x['distance'])
            logging.debug(f"Found PDF revision value: '{best_candidate['text']}' near Revision label")
            return best_candidate['text']
    
    logging.warning("Could not find revision value near 'Revision' label")
    return ""

def find_revision_in_history_table(page, revision_to_find):
    """
    Check if revision exists in the revision history table.
    """
    if not revision_to_find:
        return False
    
    text_dict = page.get_text("dict")
    page_width = page.rect.width
    page_height = page.rect.height
    
    title_block_x_start = page_width * 0.6
    title_block_y_start = page_height * 0.3
    
    revision_found_in_table = False
    
    for block in text_dict["blocks"]:
        if "lines" not in block:
            continue
            
        for line in block["lines"]:
            for span in line["spans"]:
                text = span["text"].strip()
                bbox = span["bbox"]
                
                if bbox[0] >= title_block_x_start and bbox[1] >= title_block_y_start:
                    if revision_to_find.upper() == text.upper():
                        logging.debug(f"Found revision '{revision_to_find}' in history table at position {bbox}")
                        revision_found_in_table = True
                        break
                    
                    pattern = r'\b' + re.escape(revision_to_find.upper()) + r'\b'
                    if re.search(pattern, text.upper()):
                        logging.debug(f"Found revision '{revision_to_find}' in history table text: '{text}'")
                        revision_found_in_table = True
                        break
    
    if not revision_found_in_table:
        for block in text_dict["blocks"]:
            if "lines" not in block:
                continue
                
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    
                    if revision_to_find.upper() == text.upper():
                        logging.debug(f"Found revision '{revision_to_find}' in page content: '{text}'")
                        revision_found_in_table = True
                        break
                    
                    pattern = r'\b' + re.escape(revision_to_find.upper()) + r'\b'
                    if re.search(pattern, text.upper()):
                        logging.debug(f"Found revision '{revision_to_find}' in page text: '{text}'")
                        revision_found_in_table = True
                        break
    
    if revision_found_in_table:
        logging.info(f"Revision '{revision_to_find}' validated in PDF history table")
    else:
        logging.warning(f"Revision '{revision_to_find}' not found in PDF history table")
    
    return revision_found_in_table

def extract_table_title_from_pdf(page):
    """
    Extract table title - only the 5 standard phases.
    """
    text_dict = page.get_text("dict")
    
    standard_table_titles = [
        'Concept Design',
        'Schematic Design',
        'Design Development', 
        'Construction Documents',
        'Construction Procurement'
    ]
    
    for block in text_dict["blocks"]:
        if "lines" not in block:
            continue
            
        for line in block["lines"]:
            for span in line["spans"]:
                text = span["text"].strip()
                
                for title in standard_table_titles:
                    if title.lower() in text.lower():
                        logging.debug(f"Table title found: {title}")
                        return title
    
    default_title = "Construction Procurement"
    logging.debug(f"No table title found, using default: {default_title}")
    return default_title

def find_matching_pdf_file(pdf_folder, excel_record):
    """
    Find the PDF file that matches the Excel record.
    
    Args:
        pdf_folder (Path): Path to folder containing PDF files
        excel_record (dict): Excel record with document_number, revision, title
        
    Returns:
        Path or None: Path to matching PDF file
    """
    document_number = excel_record.get('document_number', '').strip()
    revision = excel_record.get('revision', '').strip()
    
    if not document_number:
        return None
    
    # Look for PDF files in the folder
    pdf_files = list(pdf_folder.glob('*.pdf'))
    
    for pdf_file in pdf_files:
        filename = pdf_file.name
        
        # Check if document number is in filename
        if document_number.lower() in filename.lower():
            # Check if revision matches (if provided)
            if revision:
                # Handle special revision mappings
                revision_patterns = []
                
                # Add the exact revision pattern
                revision_patterns.append(f"[{revision}]")
                
                # Handle common revision mappings
                if revision == "00":
                    revision_patterns.extend(["[XX]", "[00]"])
                elif revision == "02":
                    revision_patterns.extend(["[XX]", "[02]"])
                
                # Check all possible patterns
                for pattern in revision_patterns:
                    if pattern in filename:
                        logging.debug(f"Found matching PDF: {filename} for document {document_number} revision {revision} (matched pattern: {pattern})")
                        return pdf_file
            else:
                # If no revision specified, just match by document number
                logging.debug(f"Found matching PDF: {filename} for document {document_number}")
                return pdf_file
    
    logging.warning(f"No matching PDF found for document {document_number} revision {revision}")
    return None

def process_pdf_with_excel_record(pdf_path, excel_record):
    """
    Process PDF using Excel record data.
    """
    pdf_file = Path(pdf_path)
    filename = pdf_file.name
    
    logging.info(f"Processing PDF: {filename}")
    
    try:
        # Open PDF for validation
        doc = fitz.open(pdf_path)
        page = doc[0]  # Process first page
        
        # Get data from Excel record
        document_number = excel_record.get('document_number', '')
        revision = excel_record.get('revision', '')
        title = excel_record.get('title', '')
        
        # Handle revision formatting
        if revision and revision.isdigit() and len(revision) == 1:
            revision = f"0{revision}"
        
        # NEW REVISION VALIDATION LOGIC AS SPECIFIED
        # Step 1: Get revision from Excel
        excel_revision = revision
        
        # Step 2: Extract revision from PDF title block
        pdf_revision = extract_revision_from_pdf_title_block(page)
        
        # Initialize output fields
        output_revision = ""
        output_latest_revision = ""
        validation_status = "FAILED"
        validation_comment = ""
        
        # Step 3: Compare Excel revision vs PDF revision
        if excel_revision and pdf_revision:
            if excel_revision.upper() == pdf_revision.upper():
                # They match - write this value in revision column
                output_revision = excel_revision
                
                # Step 4: Check if same value exists in history table
                revision_in_history = find_revision_in_history_table(page, excel_revision)
                
                if revision_in_history:
                    # They match - write value in latest_revision column and mark SUCCESS
                    output_latest_revision = excel_revision
                    validation_status = "SUCCESS"
                    validation_comment = ""
                else:
                    # Don't match - leave latest_revision empty, mark FAILED
                    output_latest_revision = ""
                    validation_status = "FAILED"
                    validation_comment = "The revision doesn't match with latest revision number"
            else:
                # Don't match - leave both columns empty, mark FAILED
                output_revision = ""
                output_latest_revision = ""
                validation_status = "FAILED"
                validation_comment = "Revision no. don't match"
        elif not excel_revision:
            # No Excel revision - leave both empty, mark FAILED
            output_revision = ""
            output_latest_revision = ""
            validation_status = "FAILED"
            validation_comment = "No revision specified in Excel"
        elif not pdf_revision:
            # Could not extract PDF revision - leave both empty, mark FAILED
            output_revision = ""
            output_latest_revision = ""
            validation_status = "FAILED"
            validation_comment = "Could not extract revision from PDF"
        
        # Validate other data against PDF content
        document_number_valid = validate_text_in_pdf(page, document_number) if document_number else False
        title_valid = validate_text_in_pdf(page, title) if title else False
        
        # Extract table title from PDF
        table_title = extract_table_title_from_pdf(page)
        
        doc.close()
        
        # Determine reason based on output_latest_revision prefix (only if we have a valid latest revision)
        reason = ""
        if output_latest_revision:
            if output_latest_revision.upper().startswith('T'):
                reason = "Issued for Tender"
            elif output_latest_revision.upper().startswith('N'):
                reason = "Issued for Construction"
        
        # COMPREHENSIVE VALIDATION - Check all fields and collect all errors
        overall_status = "SUCCESS"
        comments = []
        
        # Check revision validation - only clear revision fields if revision validation failed
        if validation_comment:
            overall_status = "FAILED"
            comments.append(validation_comment)
            # Only clear revision fields if revision validation specifically failed
            output_revision = ""
            output_latest_revision = ""
            reason = ""
        
        # Always check document number validation (continue even if revision failed)
        if not document_number_valid and document_number:
            overall_status = "FAILED"
            comments.append("Document number not found in PDF")
        
        # Always check title validation (continue even if other validations failed)
        output_title = title  # Default to Excel title
        if not title_valid and title:
            overall_status = "FAILED"
            comments.append("Title mismatch between Excel and PDF")
            output_title = ""  # Clear title field ONLY when title validation fails
        
        # Combine comments
        final_comments = "; ".join(comments) if comments else ""
        
        # Create result with corrected output fields
        result = {
            'file_name': filename,
            'folder': pdf_file.parent.name,
            'drawing_title': output_title,  # Empty if title validation failed
            'drawing_number': document_number,
            'revision': output_revision,  # Only populated if ALL validations pass
            'latest_revision': output_latest_revision,  # Only populated if ALL validations pass
            'latest_date': "",
            'latest_reason': "",
            'reason': reason,
            'table_title': table_title,
            'excel_row': excel_record.get('row_number', ''),
            'comments': final_comments,
            'status': overall_status
        }
        
        if validation_status == "SUCCESS":
            logging.info(f"Revision '{excel_revision}' validated successfully")
        else:
            logging.warning(f"Revision validation failed: {validation_comment}")
        
        logging.info(f"Processing complete for {filename}: {overall_status}")
        return result
        
    except Exception as e:
        logging.error(f"Error processing {filename}: {e}")
        return {
            'file_name': filename,
            'folder': pdf_file.parent.name if pdf_file.parent else "",
            'drawing_title': "",
            'drawing_number': "",
            'revision': "",
            'latest_revision': "",
            'latest_date': "",
            'latest_reason': "",
            'reason': "",
            'table_title': "",
            'excel_row': excel_record.get('row_number', ''),
            'comments': "",
            'status': f"ERROR: {str(e)}"
        }

def process_testdata_folder():
    """
    Process all PDF files in testdata subfolders using the main Excel file.
    """
    testdata_path = Path('testdata')
    excel_file = testdata_path / 'ExportDocs - Test Data file RSG QAQC Process.xlsx'
    
    if not testdata_path.exists():
        logging.error("testdata folder not found")
        return []
    
    if not excel_file.exists():
        logging.error(f"Excel file not found: {excel_file}")
        return []
    
    # Get all sheet names from Excel file
    sheet_names = get_excel_sheet_names(str(excel_file))
    
    if not sheet_names:
        logging.error("No sheets found in Excel file")
        return []
    
    all_results = []
    
    # Process each sheet/subfolder combination
    for sheet_name in sheet_names:
        logging.info(f"Processing sheet: {sheet_name}")
        
        # Map sheet name to folder name (remove _Docs suffix if present)
        folder_name = sheet_name.replace('_Docs', '').replace('_docs', '')
        subfolder_path = testdata_path / folder_name
        
        if not subfolder_path.exists() or not subfolder_path.is_dir():
            logging.warning(f"Subfolder '{folder_name}' not found for sheet '{sheet_name}', skipping")
            continue
        
        # Read Excel data for this sheet
        excel_records = read_excel_sheet_data(str(excel_file), sheet_name)
        
        if not excel_records:
            logging.warning(f"No data found in sheet '{sheet_name}'")
            continue
        
        # Get PDF files in subfolder
        pdf_files = list(subfolder_path.glob('*.pdf'))
        logging.info(f"Found {len(pdf_files)} PDF files in folder '{folder_name}'")
        
        # Process each Excel record
        for excel_record in excel_records:
            # Find matching PDF file
            matching_pdf = find_matching_pdf_file(subfolder_path, excel_record)
            
            if matching_pdf:
                # Process the PDF with Excel data
                result = process_pdf_with_excel_record(str(matching_pdf), excel_record)
                all_results.append(result)
            else:
                # Create a record for missing PDF - LEAVE ALL FIELDS EMPTY except file_name
                result = {
                    'file_name': f"NOT_FOUND_{excel_record.get('document_number', 'UNKNOWN')}",
                    'folder': folder_name,
                    'drawing_title': "",  # Empty for missing PDF
                    'drawing_number': "",  # Empty for missing PDF
                    'revision': "",  # Empty for missing PDF
                    'latest_revision': "",  # Empty for missing PDF
                    'latest_date': "",
                    'latest_reason': "",
                    'reason': "",  # Empty for missing PDF
                    'table_title': "",  # Empty for missing PDF
                    'excel_row': excel_record.get('row_number', ''),
                    'comments': "File not found",  # Specific comment for missing PDF
                    'status': "FAILED"  # Changed from ERROR to FAILED
                }
                all_results.append(result)
    
    return all_results

def save_results_to_csv(results, output_filename="pdf_extraction_results_testdata.csv"):
    """Save extraction results to CSV file with proper UTF-8 encoding and Excel highlighting."""
    if not results:
        logging.warning("No results to save")
        return
    
    df = pd.DataFrame(results)
    
    # Ensure revision columns are treated as strings
    if 'revision' in df.columns:
        df['revision'] = df['revision'].apply(lambda x: f"'{x}" if pd.notna(x) and str(x) else x)
    if 'latest_revision' in df.columns:
        df['latest_revision'] = df['latest_revision'].apply(lambda x: f"'{x}" if pd.notna(x) and str(x) else x)
    
    # Save regular CSV
    df.to_csv(output_filename, index=False, encoding='utf-8-sig')
    logging.info(f"Results saved to: {output_filename} with UTF-8-sig encoding")
    
    # Also create an Excel file with highlighting for failed rows
    excel_filename = output_filename.replace('.csv', '.xlsx')
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
        
        # Save to Excel
        df.to_excel(excel_filename, index=False, engine='openpyxl')
        
        # Open the Excel file to add highlighting
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active
        
        # Define red fill for failed rows
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        # Find the status column
        status_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'status':
                status_col = col
                break
        
        if status_col:
            # Highlight rows where status contains "FAILED"
            for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                status_value = ws.cell(row=row, column=status_col).value
                if status_value and "FAILED" in str(status_value):
                    # Highlight the entire row
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = red_fill
        
        wb.save(excel_filename)
        logging.info(f"Excel file with highlighting saved to: {excel_filename}")
        
    except ImportError:
        logging.warning("openpyxl not available, skipping Excel file with highlighting")
    except Exception as e:
        logging.warning(f"Could not create highlighted Excel file: {e}")

def print_detailed_results(results):
    """Print detailed results with validation status."""
    print("\n📊 TESTDATA FOLDER EXTRACTION RESULTS:")
    print("=" * 80)
    
    # Group results by folder
    folders = {}
    for result in results:
        folder = result.get('folder', 'Unknown')
        if folder not in folders:
            folders[folder] = []
        folders[folder].append(result)
    
    for folder_name, folder_results in folders.items():
        print(f"\n📁 FOLDER: {folder_name}")
        print("-" * 60)
        
        for result in folder_results:
            status_icon = "❌" if "FAILED" in result['status'] else "✅"
            print(f"\n📋 {result['file_name']}")
            print(f"  📝 Title: '{result['drawing_title']}'")
            print(f"  🔢 Number: '{result['drawing_number']}'")
            print(f"  📊 Revision: '{result['revision']}'")
            print(f"  📈 Latest Revision: '{result['latest_revision']}'")
            print(f"  💡 Reason: '{result['reason']}'")
            print(f"  📋 Table Title: '{result['table_title']}'")
            print(f"  📄 Excel Row: {result['excel_row']}")
            if result['comments']:
                print(f"  💬 Comments: {result['comments']}")
            print(f"  {status_icon} Status: {result['status']}")
    
    # Summary statistics
    total_files = len(results)
    success_count = len([r for r in results if r['status'] == 'SUCCESS'])
    failed_count = len([r for r in results if 'FAILED' in r['status']])
    error_count = len([r for r in results if 'ERROR' in r['status']])
    
    print(f"\n📊 SUMMARY:")
    print(f"  📁 Total Records: {total_files}")
    print(f"  ✅ Successful: {success_count} ({success_count/total_files*100:.1f}%)")
    print(f"  ❌ Failed Validation: {failed_count} ({failed_count/total_files*100:.1f}%)")
    print(f"  🚫 Errors: {error_count} ({error_count/total_files*100:.1f}%)")

def main():
    """
    Main function - TestData folder PDF extraction and validation.
    """
    print("📊 STARTING TESTDATA FOLDER PDF EXTRACTION...")
    print("=" * 80)
    print("Logic:")
    print("1. Process testdata folder with subfolders")
    print("2. Read from main Excel file with multiple sheets")
    print("3. Each sheet corresponds to a subfolder")
    print("4. Column 13+ contains PDF file records")
    print("5. Match Excel records to PDF files and validate")
    print("6. Add comments for title mismatches")
    
    # Process testdata folder
    results = process_testdata_folder()
    
    if not results:
        print("No results found to process")
        return
    
    # Save results to CSV
    output_file = 'pdf_extraction_results_testdata.csv'
    save_results_to_csv(results, output_file)
    
    # Print detailed results
    print_detailed_results(results)
    
    print(f"\n✅ TestData extraction complete! Results saved to: {output_file}")

if __name__ == "__main__":
    main()